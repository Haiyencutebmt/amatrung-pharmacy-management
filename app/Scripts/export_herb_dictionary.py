import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    if len(sys.argv) < 4:
        print("Usage: export_herb_dictionary.py <input_json_path> <template_xlsx_path> <output_xlsx_path>")
        sys.exit(1)
        
    input_json_path = sys.argv[1]
    template_xlsx_path = sys.argv[2]
    output_xlsx_path = sys.argv[3]
    
    try:
        # Đọc dữ liệu JSON
        with open(input_json_path, 'r', encoding='utf-8') as f:
            entries = json.load(f)
            
        # Mở file excel mẫu
        wb = openpyxl.load_workbook(template_xlsx_path)
        ws = wb.active
        
        # Đảm bảo hiển thị lưới (gridlines)
        ws.views.sheetView[0].showGridLines = True
        
        # Xóa các dòng dữ liệu mẫu (từ dòng 7 trở đi)
        if ws.max_row >= 7:
            ws.delete_rows(7, ws.max_row - 6)
            
        # Định nghĩa các style cơ bản
        font_family = "Segoe UI"
        font_data = Font(name=font_family, size=10, color="000000")
        
        # Màu nền xen kẽ (Row 7, 9, 11... là F8FAFC, Row 8, 10, 12... là FFFFFF)
        fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Khung viền mỏng màu xám CBD5E1
        thin_side = Side(style='thin', color='CBD5E1')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Căn lề trái, dọc giữa, wrap text
        alignment_data = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        current_row = 7
        for idx, entry in enumerate(entries):
            row_values = [
                entry.get('name', '') or '',
                entry.get('scientific_name', '') or '',
                entry.get('other_names', '') or '',
                entry.get('family', '') or '',
                entry.get('plant_part', '') or '',
                entry.get('properties', '') or '',
                entry.get('basic_info', '') or '',
                entry.get('effects', '') or '',
                entry.get('usage_notes', '') or '',
                entry.get('safety_warning', '') or '',
                entry.get('status', '') or ''
            ]
            
            # Chọn màu nền xen kẽ
            current_fill = fill_even if idx % 2 == 0 else fill_odd
            
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = border_all
                cell.alignment = alignment_data
                cell.fill = current_fill
                
            current_row += 1
            
        # Lưu file kết quả
        wb.save(output_xlsx_path)
        print("Success")
        
    except Exception as e:
        import traceback
        print(f"Error: {e}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
