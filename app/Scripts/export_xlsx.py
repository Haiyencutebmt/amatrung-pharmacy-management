import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) < 3:
        print("Usage: export_xlsx.py <input_json_path> <output_xlsx_path>")
        sys.exit(1)
        
    input_json_path = sys.argv[1]
    output_xlsx_path = sys.argv[2]
    
    try:
        with open(input_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách"
        
        # Đảm bảo hiển thị lưới (gridlines)
        ws.views.sheetView[0].showGridLines = True
        
        # Định nghĩa các style cơ bản
        font_family = "Segoe UI"
        font_title = Font(name=font_family, size=16, bold=True, color="1B5E20") # Xanh lá đậm
        font_sub = Font(name=font_family, size=10, italic=True, color="555555")
        font_clinic_name = Font(name=font_family, size=11, bold=True, color="000000")
        font_clinic_info = Font(name=font_family, size=9, color="333333")
        font_header = Font(name=font_family, size=10, bold=True, color="000000")
        font_data = Font(name=font_family, size=10, color="000000")
        
        fill_header = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Xanh lá nhạt nhã nhặn
        
        thin_side = Side(style='thin', color='B0BEC5')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # 1. Ghi thông tin phòng khám ở góc trên bên trái
        clinic = data.get("clinic_info", {})
        ws["A1"] = clinic.get("name", "NHÀ THUỐC AMATRUNG").upper()
        ws["A1"].font = font_clinic_name
        
        ws["A2"] = f"Địa chỉ: {clinic.get('address', '')}"
        ws["A2"].font = font_clinic_info
        
        ws["A3"] = f"Hotline: {clinic.get('phone', '')} | MST: {clinic.get('mst', '066070008130')}"
        ws["A3"].font = font_clinic_info
        
        ws["A4"] = f"Bác sĩ phụ trách: {clinic.get('doctor', 'BS. Y Hiếu Niê')}"
        ws["A4"].font = font_clinic_info
        
        # 2. Ghi Tiêu đề chính ở dòng 6
        title_text = data.get("title", "DANH SÁCH")
        ws["A6"] = title_text
        ws["A6"].font = font_title
        
        # Căn lề trái hoặc merge dòng tiêu đề tùy ý, ở đây để ở A6 và không merge để tránh rắc rối tự động co giãn cột.
        # Thêm thông tin ngày xuất bản
        import datetime
        now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A7"] = f"Ngày tải về: {now_str} | Tổng số dòng: {len(data.get('rows', []))}"
        ws["A7"].font = font_sub
        
        # 3. Bắt đầu ghi bảng từ dòng 9
        start_row = 9
        headers = data.get("headers", [])
        alignments = data.get("alignments", [])
        
        # Ghi Header
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        # Ghi Rows
        current_row = start_row + 1
        for row_data in data.get("rows", []):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = font_data
                cell.border = border_all
                
                # Căn lề
                align = "left"
                if col_idx - 1 < len(alignments):
                    align = alignments[col_idx - 1]
                
                cell.alignment = Alignment(horizontal=align, vertical="center")
                
                # Định dạng số cho cột số lượng tồn kho (nếu giá trị là float/int)
                if isinstance(val, (int, float)) and col_idx > 1:
                    cell.number_format = '#,##0'
                    
            current_row += 1
            
        # 4. Tự co giãn cột
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Tính độ rộng dựa trên các dòng dữ liệu (từ dòng 9 trở đi để không bị ảnh hưởng bởi tiêu đề phòng khám)
            for cell in col[8:]: # Từ dòng 9 (0-indexed thì index 8 là dòng 9)
                if cell.value:
                    val_str = str(cell.value)
                    # Xử lý tiếng Việt có dấu làm tăng chiều dài hiển thị
                    max_len = max(max_len, len(val_str))
            
            # Đảm bảo độ rộng tối thiểu và cộng thêm padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Lưu file
        wb.save(output_xlsx_path)
        print("Success")
    except Exception as e:
        import traceback
        print(f"Error: {e}")
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
