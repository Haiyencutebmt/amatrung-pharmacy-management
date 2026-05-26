import os

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUTPUT_PATH = os.path.join(ROOT_DIR, "public", "templates", "mau_import_tu_dien_thuoc_nam.xlsx")


def style_cell(cell, font=None, fill=None, border=None, alignment=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment


def generate_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhap tu dien"
    ws.sheet_view.showGridLines = False

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="064E3B")
    subtitle_font = Font(name=font_family, size=10, color="475569")
    note_font = Font(name=font_family, size=9, italic=True, color="92400E")
    header_font = Font(name=font_family, size=10, bold=True, color="0F172A")
    required_header_font = Font(name=font_family, size=10, bold=True, color="7F1D1D")
    data_font = Font(name=font_family, size=10, color="1E293B")
    muted_font = Font(name=font_family, size=9, color="64748B")

    title_fill = PatternFill("solid", fgColor="DCFCE7")
    note_fill = PatternFill("solid", fgColor="FFFBEB")
    header_fill = PatternFill("solid", fgColor="E0F2FE")
    required_fill = PatternFill("solid", fgColor="FEE2E2")
    sample_fill = PatternFill("solid", fgColor="F8FAFC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    thin_green = Side(style="thin", color="86EFAC")
    thin_gray = Side(style="thin", color="CBD5E1")
    header_border = Border(left=thin_green, right=thin_green, top=thin_green, bottom=thin_green)
    data_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A1:K1")
    ws["A1"] = "MẪU NHẬP TỪ ĐIỂN THUỐC NAM"
    style_cell(ws["A1"], title_font, title_fill, None, center)

    ws.merge_cells("A2:K2")
    ws["A2"] = "Nhập dữ liệu từ dòng tiêu đề bên dưới. Hình ảnh không nhập qua Excel, có thể bổ sung sau ở nút Sửa từng mục."
    style_cell(ws["A2"], subtitle_font, None, None, center)

    ws.merge_cells("A4:K4")
    ws["A4"] = "Cột có dấu * là bắt buộc. Không đổi tên các cột. Có thể xóa 2 dòng ví dụ trước khi nhập dữ liệu thật."
    style_cell(ws["A4"], note_font, note_fill, data_border, center)

    headers = [
        ("Tên thuốc nam *", True, "Bắt buộc. Đây là tên hiển thị chính trên website."),
        ("Tên khoa học", False, "Tên Latin hoặc tên khoa học, nếu có."),
        ("Tên gọi khác", False, "Các tên dân gian, cách nhau bằng dấu phẩy."),
        ("Họ thực vật", False, "VD: Fabaceae, Apiaceae..."),
        ("Bộ phận dùng", False, "VD: toàn cây, lá, rễ, thân, hoa..."),
        ("Tính vị / đặc điểm", False, "Mô tả ngắn để hiển thị nhanh trên danh sách."),
        ("Thông tin cơ bản *", True, "Bắt buộc. Mô tả nhận diện và thông tin nền."),
        ("Tác dụng *", True, "Bắt buộc. Ghi tác dụng tham khảo, không ghi như chỉ định thay thăm khám."),
        ("Lưu ý khi sử dụng", False, "Chống chỉ định, đối tượng cần thận trọng, cách dùng an toàn."),
        ("Khuyến cáo an toàn", False, "Nếu bỏ trống, hệ thống tự thêm khuyến cáo không tự ý sử dụng."),
        ("Trạng thái", False, "Chọn Đã xuất bản hoặc Bản nháp."),
    ]

    header_row = 6
    for col_idx, (label, required, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.comment = Comment(comment, "AmaTrung")
        style_cell(
            cell,
            required_header_font if required else header_font,
            required_fill if required else header_fill,
            header_border,
            center,
        )

    sample_rows = [
        [
            "Kim Tiền Thảo",
            "Desmodium styracifolium (Osbeck) Merr.",
            "Mắt trâu, đồng tiền lông",
            "Fabaceae",
            "Toàn cây",
            "Vị ngọt, tính mát",
            "Cây thuốc nam thường được ghi nhận trong tài liệu y học cổ truyền. Cần định danh đúng cây trước khi sử dụng.",
            "Hỗ trợ lợi tiểu, thanh nhiệt theo chỉ định của thầy thuốc.",
            "Không dùng thay thế thuốc điều trị. Người có bệnh nền, phụ nữ có thai hoặc đang dùng thuốc khác cần hỏi thầy thuốc.",
            "Không nên tự ý sử dụng thuốc nam khi chưa được thăm khám và tư vấn phù hợp.",
            "Đã xuất bản",
        ],
        [
            "Rau Má",
            "Centella asiatica",
            "Tích tuyết thảo",
            "Apiaceae",
            "Toàn cây",
            "Vị đắng nhẹ, tính mát",
            "Cây thảo mọc bò, thường gặp ở vùng ẩm. Thông tin chỉ dùng để tham khảo nhận diện.",
            "Thường được ghi nhận trong hỗ trợ thanh nhiệt theo chỉ định.",
            "Không tự phối hợp nhiều vị thuốc. Ngưng sử dụng nếu có biểu hiện bất thường và liên hệ thầy thuốc.",
            "",
            "Bản nháp",
        ],
    ]

    for row_idx, row in enumerate(sample_rows, start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_cell(cell, data_font, sample_fill if row_idx % 2 else white_fill, data_border, left)
        ws.row_dimensions[row_idx].height = 76

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = "A6:K8"

    column_widths = {
        "A": 24,
        "B": 30,
        "C": 28,
        "D": 18,
        "E": 18,
        "F": 24,
        "G": 46,
        "H": 42,
        "I": 48,
        "J": 46,
        "K": 16,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[6].height = 44

    status_validation = DataValidation(type="list", formula1='"Đã xuất bản,Bản nháp"', allow_blank=True)
    ws.add_data_validation(status_validation)
    status_validation.add("K7:K5000")

    for row in range(9, 5001):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            style_cell(cell, muted_font, None, data_border, left)

    guide = wb.create_sheet("Huong dan")
    guide.sheet_view.showGridLines = False
    guide.merge_cells("A1:D1")
    guide["A1"] = "HƯỚNG DẪN NHẬP FILE TỪ ĐIỂN"
    style_cell(guide["A1"], title_font, title_fill, None, center)

    guide_headers = ["Cột", "Bắt buộc", "Gợi ý nhập", "Ví dụ"]
    for col_idx, label in enumerate(guide_headers, start=1):
        cell = guide.cell(row=3, column=col_idx, value=label)
        style_cell(cell, header_font, header_fill, header_border, center)

    guide_rows = [
        ["Tên thuốc nam", "Có", "Tên ngắn, rõ, đúng chính tả.", "Kim Tiền Thảo"],
        ["Thông tin cơ bản", "Có", "Mô tả nhận diện, nguồn gốc, bộ phận dùng.", "Cây thuốc nam thường dùng trong YHCT..."],
        ["Tác dụng", "Có", "Ghi tác dụng tham khảo, tránh khẳng định thay thế điều trị.", "Hỗ trợ lợi tiểu theo chỉ định..."],
        ["Khuyến cáo an toàn", "Không", "Có thể để trống để hệ thống tự thêm cảnh báo an toàn.", "Không nên tự ý sử dụng..."],
        ["Trạng thái", "Không", "Đã xuất bản sẽ hiển thị trên website; Bản nháp chỉ lưu trong admin.", "Đã xuất bản"],
    ]

    for row_idx, row in enumerate(guide_rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = guide.cell(row=row_idx, column=col_idx, value=value)
            style_cell(cell, data_font, white_fill, data_border, left)
        guide.row_dimensions[row_idx].height = 38

    for letter, width in {"A": 24, "B": 14, "C": 58, "D": 34}.items():
        guide.column_dimensions[letter].width = width

    guide.row_dimensions[1].height = 30
    guide.row_dimensions[3].height = 30

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.active = 0
    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    generate_template()
