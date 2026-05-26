import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Duoc Lieu"
    
    # Hiện lưới
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    font_eng = Font(name=font_family, size=9, bold=False, color="7F8C8D") # Chữ xám nhỏ cho key tiếng Anh
    font_vie = Font(name=font_family, size=10, bold=True, color="2C3E50") # Chữ đậm cho tiếng Việt
    font_data = Font(name=font_family, size=10, color="000000")
    
    fill_vie = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Màu xanh lá nhạt trang nhã
    
    thin_side = Side(style='thin', color='BDC3C7')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Header tiếng Anh (Dòng 1)
    headers_eng = ['name', 'category', 'usage_type', 'unit', 'stock_quantity', 'expiry_date', 'status', 'warning_note', 'description']
    # Header tiếng Việt (Dòng 2)
    headers_vie = ['Tên Dược Liệu (*)', 'Phân Loại', 'Cách Dùng', 'Đơn Vị Tính', 'Số Lượng Tồn', 'Hạn Sử Dụng (YYYY-MM-DD)', 'Trạng Thái', 'Ghi Chú Cảnh Báo', 'Mô Tả Chi Tiết']
    
    # Ghi dòng 1
    for col_idx, val in enumerate(headers_eng, start=1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font = font_eng
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
        
    # Ghi dòng 2
    for col_idx, val in enumerate(headers_vie, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = font_vie
        cell.fill = fill_vie
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        
    # Dữ liệu mẫu (Dòng 3 & 4)
    data_rows = [
        ['Cam thảo bắc', 'Dược liệu bốc thuốc', 'Sắc', 'g', 1500, '2027-12-31', 'Đang dùng', 'Tránh ẩm mốc', 'Bổ tỳ vị, nhuận phế, thanh nhiệt giải độc.'],
        ['Đương quy', 'Dược liệu bốc thuốc', 'Sắc', 'g', 800, '2026-10-15', 'Đang dùng', 'Dễ bị ẩm và mốc mọt', 'Bổ huyết, hoạt huyết, nhuận tràng thông tiện.']
    ]
    
    for row_idx, row_data in enumerate(data_rows, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_all
            
            # Căn lề hợp lý
            align = "left"
            if col_idx in [4, 6, 7]: # Đơn vị, Hạn sử dụng, Trạng thái
                align = "center"
            elif col_idx == 5: # Số lượng tồn
                align = "right"
                cell.number_format = '#,##0'
                
            cell.alignment = Alignment(horizontal=align, vertical="center")
            
    # Co giãn cột tự động
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Đếm độ dài chuỗi
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 28
    
    wb.save('public/templates/mau_import_duoc_lieu.xlsx')
    print("Template generated successfully!")

if __name__ == "__main__":
    generate_template()
